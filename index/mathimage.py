import numpy as np
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D  # 只需导入即可激活 3D
from matplotlib.patches import Circle
from mpl_toolkits.mplot3d import art3d
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
import matplotlib
import io
import base64


matplotlib.use('Agg')  # 使用非交互式后端
# 配置 matplotlib 中文显示
plt.rcParams['font.sans-serif'] = ['SimHei']  # 使用黑体字体
plt.rcParams['axes.unicode_minus'] = False  # 解决负号显示为方块的问题
plt.rcParams['toolbar'] = 'None'# 关闭工具栏


def status(data,color="#FFFFFFFF"):
    """
    绘制过去7天的任务完成情况
    :param data: 字典，键为日期，值为元组(完成,未完成)
    :return: base64编码的图片字符串
    """
    # 将字典按日期排序，方便横轴展示
    dates = sorted(data.keys())
    complete = [data[d][0] for d in dates]
    incomplete = [data[d][1] for d in dates]

    x = np.arange(len(dates))  # 横轴刻度位置

    fig, ax = plt.subplots(figsize=(6, 4))  # 设置较小的图表尺寸
    fig.patch.set_facecolor(color)   # 画布背景
    ax.set_facecolor(color)          # 坐标轴背景

    # 画堆叠柱：先完成，后未完成
    bars1 = ax.bar(x, complete, color="#5238AF", label='完成', alpha=0.8)
    bars2 = ax.bar(x, incomplete, bottom=complete, color="#0CAB91", label='未完成', alpha=0.8)

    # 在柱顶标注总数，只在有数据的柱子顶部显示
    for i, (c, inc) in enumerate(zip(complete, incomplete)):
        total = c + inc
        if total > 0:  # 只在有数据时显示标签
            ax.annotate(str(total), xy=(i, total),
                        xytext=(0, 5), textcoords="offset points",
                        ha='center', va='bottom', fontsize=10, fontweight='bold')

    ax.set_title('过去7天的任务完成情况', fontsize=14, fontweight='bold', pad=20)
    ax.set_ylabel('任务数量', fontsize=12)
    ax.set_xlabel('日期', fontsize=12)
    ax.set_xticks(x)
    ax.set_xticklabels(dates, rotation=40, ha='right')
    ax.legend(loc='best')  # 自动选择最佳图例位置
    
    # 让Y轴自动适应数据大小，添加一些padding使图表更美观
    max_total = max(c + inc for c, inc in zip(complete, incomplete))
    ax.set_ylim(0, max_total * 1.1)  # 增加10%的上边距
    
    plt.tight_layout()
    
    # 将图片转换为base64编码
    buffer = io.BytesIO()
    plt.savefig(buffer, format='png', bbox_inches='tight', dpi=150, facecolor=color)
    buffer.seek(0)
    image_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
    buffer.close()
    plt.close()
    
    return f"data:image/png;base64,{image_base64}"

def status2(data, color="#FFFFFFFF"):
    """
    绘制任务完成情况的圆饼图
    :param data: 列表，[完成, 未完成]
    :return: base64编码的图片字符串
    """
    labels = ['完成', '未完成']
    sizes = data
    colors = ['#5238AF', '#0CAB91']

    fig, ax = plt.subplots(figsize=(5, 4))  # 设置较小的图表尺寸
    fig.patch.set_facecolor(color)  # 画布背景
    ax.set_facecolor(color)         # 坐标轴背景

    wedges, texts, autotexts = ax.pie(sizes,
           labels=labels,
           colors=colors,
           autopct='%1.1f%%',
           shadow=False,  # 关闭阴影，避免重影
           startangle=90,
           textprops={'fontsize': 10, 'color': '#333333'},
           wedgeprops={'linewidth': 1, 'edgecolor': 'white'})
    
    # 优化标签样式
    for autotext in autotexts:
        autotext.set_color('white')
        autotext.set_fontweight('bold')
        autotext.set_fontsize(10)
    ax.axis('equal')  # 保证圆饼图是圆形
    ax.set_title('过去30天的任务完成情况', fontsize=14, fontweight='bold', pad=20)
    
    # 添加图例
    ax.legend(wedges, labels, title="状态", loc="center left", bbox_to_anchor=(1, 0, 0.5, 1))
    
    plt.tight_layout()
    
    # 将图片转换为base64编码
    buffer = io.BytesIO()
    plt.savefig(buffer, format='png', bbox_inches='tight', dpi=150, facecolor=color)
    buffer.seek(0)
    image_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
    buffer.close()
    plt.close()
    
    return f"data:image/png;base64,{image_base64}"

def create_task_excel(tasks, filename="task_list.xlsx", save_to_memory=False):
    """
    创建任务清单 Excel 文件
    :param tasks: 任务列表
    :param filename: 文件名
    :param save_to_memory: 是否保存到内存（True）还是保存到文件（False）
    :return: 如果 save_to_memory=True，返回字节数据；否则返回文件路径
    """
    # 创建工作簿和工作表
    wb = Workbook()
    ws = wb.active
    ws.title = "任务清单"
    
    # 定义表头
    headers = ["序号", "任务描述", "状态"]
    ws.append(headers)
    
    # 设置列宽（根据内容调整，防止显示不全）
    column_widths = [8, 50, 15]  # 加宽任务描述列，缩窄状态列
    for col, width in enumerate(column_widths, 1):  # 列号从1开始
        ws.column_dimensions[chr(64 + col)].width = width  # A=65, B=66...
    
    # 设置表头样式
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 设置表头行高
    ws.row_dimensions[1].height = 25
    
    # 导入任务数据
    for i, task in enumerate(tasks, 1):
        row = [
            i,
            task["description"],
            task["status"]
        ]
        ws.append(row)
        
        # 设置任务描述自动换行（第2列是任务描述）
        ws.cell(row=i+1, column=2).alignment = Alignment(wrap_text=True, vertical="center")
        
    
    if save_to_memory:
        # 保存到内存
        from io import BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    else:
        # 保存到文件
        wb.save(filename)
        return filename


if __name__ == "__main__":
    pass